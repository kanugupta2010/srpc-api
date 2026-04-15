"""
services/purchase_import_service.py
SRPC Enterprises Private Limited

Parses Busy 21 Purchase Register XLSX/CSV export.

Busy 21 purchase export columns:
  Date | Vch Type | Vch/Bill No | Particulars | Alias | Item Details |
  Qty. | Unit | Price | Amount

  Vch Type = 'Purc' → purchase
  Vch Type = 'PrRt' → purchase_return

Key differences from sales format:
  - No Party Mobile or Referred By columns
  - Particulars = supplier name
  - Price = unit price EXCLUDING tax
  - PrRt: qty and amount are negative
  - Each PrRt row has its own date (unlike sales which only have
    date on the first/header row)
  - Bill number may be NULL even on header rows
"""

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import openpyxl

log = logging.getLogger(__name__)

# Column names
COL_DATE         = "Date"
COL_VCH_TYPE     = "Vch Type"
COL_BILL_NO      = "Vch/Bill No"
COL_PARTICULARS  = "Particulars"
COL_ALIAS        = "Alias"
COL_ITEM_DETAILS = "Item Details"
COL_QTY          = "Qty."
COL_QTY_ALT      = "Qty"
COL_UNIT         = "Unit"
COL_PRICE        = "Price"
COL_AMOUNT       = "Amount"

REQUIRED_COLS = {COL_ALIAS, COL_AMOUNT}

VCH_PURCHASE        = "Purc"
VCH_PURCHASE_RETURN = "PrRt"

INV_PURCHASE        = "purchase"
INV_PURCHASE_RETURN = "purchase_return"


def get_financial_year(d) -> str:
    """Indian FY runs April–March. Returns e.g. '2526' for FY 2025-26."""
    if d is None:
        from datetime import date
        d = date.today()
    year, month = d.year, d.month
    if month >= 4:
        return f"{str(year)[2:]}{str(year+1)[2:]}"
    return f"{str(year-1)[2:]}{str(year)[2:]}" 


@dataclass
class ParsedPurchaseLine:
    item_code:   str
    item_name:   str
    quantity:    float
    unit:        str
    unit_price_exc: float
    line_amount_exc: float


@dataclass
class ParsedPurchaseInvoice:
    invoice_date:  Optional[date]
    bill_number:   Optional[str]
    supplier_name: str
    invoice_type:  str   # purchase | purchase_return
    financial_year: str = ""
    lines:         list = field(default_factory=list)

    @property
    def gross_amount_exc(self) -> float:
        return sum(abs(l.line_amount_exc) for l in self.lines)


# ---------------------------------------------------------------------------
# Value parsers
# ---------------------------------------------------------------------------

def _parse_date(raw) -> Optional[date]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    v = str(raw).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(v, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(raw) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    v = str(raw).strip().replace(",", "")
    try:
        return float(v)
    except ValueError:
        return 0.0


def _to_str(raw) -> str:
    return "" if raw is None else str(raw).strip()


def _get_qty(row: dict) -> float:
    return _parse_float(row.get(COL_QTY) or row.get(COL_QTY_ALT))


# ---------------------------------------------------------------------------
# Row iterators
# ---------------------------------------------------------------------------

def _iter_rows_xlsx(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the header row — it contains 'Date' and 'Vch Type'
    # Busy 21 exports may have 1-3 title rows before the actual header
    header_idx = None
    for i, row in enumerate(all_rows):
        row_vals = [_to_str(v) for v in row]
        if "Date" in row_vals and "Vch Type" in row_vals:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Could not find header row with 'Date' and 'Vch Type' columns.")

    headers = [_to_str(h) for h in all_rows[header_idx]]
    missing = REQUIRED_COLS - set(headers)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}. Found: {headers}")

    for row in all_rows[header_idx + 1:]:
        yield dict(zip(headers, row))


def _iter_rows_csv(content: bytes):
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    reader.fieldnames = [h.strip() for h in (reader.fieldnames or [])]
    for row in reader:
        yield {k.strip(): v.strip() for k, v in row.items()}


# ---------------------------------------------------------------------------
# Core parser
# ---------------------------------------------------------------------------

def _parse_rows(row_iter) -> tuple[list[ParsedPurchaseInvoice], dict]:
    invoices: list[ParsedPurchaseInvoice] = []
    current: Optional[ParsedPurchaseInvoice] = None
    stats = {"total_rows": 0, "blank_rows": 0, "invoice_rows": 0, "line_rows": 0}

    for row in row_iter:
        stats["total_rows"] += 1

        vch_type    = _to_str(row.get(COL_VCH_TYPE, ""))
        alias       = _to_str(row.get(COL_ALIAS, ""))
        bill_no     = _to_str(row.get(COL_BILL_NO, "")) or None
        particulars = _to_str(row.get(COL_PARTICULARS, ""))
        amount_raw  = row.get(COL_AMOUNT)
        date_raw    = row.get(COL_DATE)

        # Skip blank rows
        if not alias and (amount_raw is None or _to_str(amount_raw) == ""):
            stats["blank_rows"] += 1
            continue

        # Skip header row if present in data
        if vch_type == "Vch Type":
            continue

        # Determine if this row starts a new invoice
        # A new invoice starts when:
        #   - vch_type is Purc or PrRt AND date is present
        #   - OR vch_type is Purc or PrRt AND particulars is present (new supplier)
        is_new_invoice = vch_type in (VCH_PURCHASE, VCH_PURCHASE_RETURN) and (
            date_raw is not None or particulars
        )

        if is_new_invoice:
            stats["invoice_rows"] += 1
            invoice_type = INV_PURCHASE_RETURN if vch_type == VCH_PURCHASE_RETURN else INV_PURCHASE
            current = ParsedPurchaseInvoice(
                invoice_date  = _parse_date(date_raw),
                bill_number   = bill_no,
                supplier_name = particulars,
                invoice_type   = invoice_type,
                financial_year = get_financial_year(_parse_date(date_raw)),
            )
            invoices.append(current)

        # Add line item if alias exists
        if alias and current is not None:
            stats["line_rows"] += 1
            qty    = _get_qty(row)
            price  = _parse_float(row.get(COL_PRICE))
            amount = _parse_float(row.get(COL_AMOUNT))

            current.lines.append(ParsedPurchaseLine(
                item_code       = alias[:100],
                item_name       = _to_str(row.get(COL_ITEM_DETAILS, ""))[:255],
                quantity        = qty,
                unit            = _to_str(row.get(COL_UNIT, ""))[:20],
                unit_price_exc  = price,
                line_amount_exc = amount,
            ))

    return invoices, stats


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_purchase_file(content: bytes, filename: str) -> tuple[list[ParsedPurchaseInvoice], dict]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    row_iter = _iter_rows_xlsx(content) if ext == "xlsx" else _iter_rows_csv(content)
    return _parse_rows(row_iter)