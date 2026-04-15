"""
services/import_service.py
SRPC Enterprises Private Limited — Saraswati Loyalty Program

Invoice parser — reads Busy 21 XLSX or CSV export.

Two separate identifiers per invoice:
  invoice_type  → sale | sale_return
  customer_type → contractor_direct | contractor_referred | walk_in

Contractor matching — uses contractor_code only:
  Referred By field → matched directly against contractors.contractor_code
  Particulars field → matched directly against contractors.contractor_code

Busy 21 export columns:
  Date | Vch Type | Vch/Bill No | Particulars | Alias | Item Details |
  Qty. | Unit | Price | Amount | Party Name | Party Mobile | Referred By
"""

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import openpyxl

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column constants
# ---------------------------------------------------------------------------
COL_DATE         = "Date"
COL_VCH_TYPE     = "Vch Type"
COL_BILL_NO      = "Vch/Bill No"
COL_PARTICULARS  = "Particulars"
COL_ALIAS        = "Alias"
COL_ITEM_DETAILS = "Item Details"
COL_QTY          = "Qty."
COL_QTY_ALT      = "Qty"
COL_UNIT         = "Unit"
COL_PRICE        = "Price"
COL_AMOUNT       = "Amount"
COL_PARTY_NAME   = "Party Name"
COL_PARTY_MOBILE = "Party Mobile"
COL_REFERRED_BY  = "Referred By"

REQUIRED_COLS = {COL_BILL_NO, COL_ALIAS, COL_AMOUNT}

# Vch Type values from Busy 21
VCH_SALE   = "Sale"
VCH_RETURN = "SlRt"

# invoice_type values
INV_SALE        = "sale"
INV_SALE_RETURN = "sale_return"

# customer_type values
CUST_CONTRACTOR_DIRECT   = "contractor_direct"
CUST_CONTRACTOR_REFERRED = "contractor_referred"
CUST_WALK_IN             = "walk_in"


def get_financial_year(d) -> str:
    """Indian FY runs April–March. Returns e.g. '2526' for FY 2025-26."""
    if d is None:
        from datetime import date
        d = date.today()
    year, month = d.year, d.month
    if month >= 4:
        return f"{str(year)[2:]}{str(year+1)[2:]}"
    return f"{str(year-1)[2:]}{str(year)[2:]}" 


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class ParsedLine:
    item_code:   str
    item_name:   str
    quantity:    float
    unit:        str
    unit_price:  float
    line_amount: float


@dataclass
class ParsedInvoice:
    invoice_date:    Optional[date]
    bill_number:     str
    invoice_type:    str            # sale | sale_return
    customer_type:   str            # contractor_direct | contractor_referred | walk_in
    particulars:     str            # Busy 21 account name — used for contractor_code match
    party_name:      str
    party_mobile:    str
    referred_by_raw: str            # Raw Referred By value — matched as contractor_code
    contractor_id:   Optional[int] = None
    financial_year:  str = ""
    lines:           list = field(default_factory=list)

    @property
    def is_return(self) -> bool:
        return self.invoice_type == INV_SALE_RETURN

    @property
    def gross_amount(self) -> float:
        return sum(l.line_amount for l in self.lines)


# ---------------------------------------------------------------------------
# Value parsers
# ---------------------------------------------------------------------------

def _parse_date(raw) -> Optional[date]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    v = str(raw).strip()
    if not v:
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y",
                "%d %b %Y", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(v, fmt).date()
        except ValueError:
            continue
    log.warning("Could not parse date '%s'", raw)
    return None


def _parse_float(raw) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    v = str(raw).strip().replace(",", "")
    if not v:
        return 0.0
    try:
        return float(v)
    except ValueError:
        return 0.0


def _to_str(raw) -> str:
    return "" if raw is None else str(raw).strip()


def _get_qty(row: dict) -> float:
    return _parse_float(row.get(COL_QTY) or row.get(COL_QTY_ALT))


# ---------------------------------------------------------------------------
# Row iterators
# ---------------------------------------------------------------------------

def _iter_rows_xlsx(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the header row — it contains 'Date' and 'Vch Type'
    # Busy 21 exports may have 1-3 title rows before the actual header
    header_idx = None
    for i, row in enumerate(all_rows):
        row_vals = [_to_str(v) for v in row]
        if "Date" in row_vals and "Vch Type" in row_vals:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Could not find header row with 'Date' and 'Vch Type' columns.")

    headers = [_to_str(h) for h in all_rows[header_idx]]
    missing = REQUIRED_COLS - set(headers)
    if missing:
        raise ValueError(f"File missing required columns: {sorted(missing)}. Found: {headers}")

    for row in all_rows[header_idx + 1:]:
        yield dict(zip(headers, row))


def _iter_rows_csv(content: bytes):
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    reader.fieldnames = [h.strip() for h in (reader.fieldnames or [])]
    missing = REQUIRED_COLS - set(reader.fieldnames)
    if missing:
        raise ValueError(f"File missing required columns: {sorted(missing)}. Found: {reader.fieldnames}")
    for row in reader:
        yield {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}


# ---------------------------------------------------------------------------
# Core parser
# ---------------------------------------------------------------------------

def _parse_rows(row_iter) -> tuple[list[ParsedInvoice], dict]:
    invoices: list[ParsedInvoice] = []
    current: Optional[ParsedInvoice] = None
    stats = {"total_rows": 0, "blank_rows": 0, "line_rows": 0, "header_rows": 0}

    for row in row_iter:
        stats["total_rows"] += 1

        bill_no    = _to_str(row.get(COL_BILL_NO, ""))
        alias      = _to_str(row.get(COL_ALIAS, ""))
        amount_raw = row.get(COL_AMOUNT)
        vch_type   = _to_str(row.get(COL_VCH_TYPE, ""))

        # Skip completely blank rows
        if not bill_no and not alias and (amount_raw is None or _to_str(amount_raw) == ""):
            stats["blank_rows"] += 1
            continue

        # New invoice header row
        if bill_no:
            stats["header_rows"] += 1
            particulars     = _to_str(row.get(COL_PARTICULARS, ""))
            referred_by_raw = _to_str(row.get(COL_REFERRED_BY, ""))
            party_mobile    = _to_str(row.get(COL_PARTY_MOBILE, ""))

            # invoice_type — based purely on Vch Type
            invoice_type = INV_SALE_RETURN if vch_type == VCH_RETURN else INV_SALE

            # customer_type — referred_by takes priority, then particulars
            if referred_by_raw:
                customer_type = CUST_CONTRACTOR_REFERRED
            elif particulars.lower() not in ("cash", ""):
                customer_type = CUST_CONTRACTOR_DIRECT
            else:
                customer_type = CUST_WALK_IN

            current = ParsedInvoice(
                invoice_date    = _parse_date(row.get(COL_DATE)),
                bill_number     = bill_no,
                invoice_type    = invoice_type,
                customer_type   = customer_type,
                particulars     = particulars,
                party_name      = _to_str(row.get(COL_PARTY_NAME, "")),
                party_mobile    = party_mobile,
                referred_by_raw = referred_by_raw,
                contractor_id   = None,
                financial_year  = get_financial_year(_parse_date(row.get(COL_DATE))),
            )
            invoices.append(current)

        # Line item row
        if alias and current is not None:
            stats["line_rows"] += 1
            current.lines.append(ParsedLine(
                item_code   = alias[:100],
                item_name   = _to_str(row.get(COL_ITEM_DETAILS, ""))[:255],
                quantity    = _get_qty(row),
                unit        = _to_str(row.get(COL_UNIT, ""))[:20],
                unit_price  = _parse_float(row.get(COL_PRICE)),
                line_amount = _parse_float(row.get(COL_AMOUNT)),
            ))

    return invoices, stats


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_file(content: bytes, filename: str) -> tuple[list[ParsedInvoice], dict]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    row_iter = _iter_rows_xlsx(content) if ext == "xlsx" else _iter_rows_csv(content)
    return _parse_rows(row_iter)


# ---------------------------------------------------------------------------
# Contractor resolver — matches by contractor_code only
# ---------------------------------------------------------------------------

def resolve_contractors(invoices: list[ParsedInvoice], db_conn) -> None:
    """
    Resolves contractor_id for each invoice using contractor_code matching only.

    contractor_referred → Referred By field = contractor_code
    contractor_direct   → Particulars field = contractor_code
    walk_in             → no contractor match attempted
    """
    cursor = db_conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT id, contractor_code, status FROM contractors WHERE is_active = 1"
    )
    rows = cursor.fetchall()

    # Build lookup: contractor_code → contractor row
    code_map: dict[str, dict] = {}
    for r in rows:
        if r["contractor_code"]:
            code_map[r["contractor_code"].strip()] = r

    cursor.close()

    for inv in invoices:

        if inv.customer_type == CUST_CONTRACTOR_REFERRED:
            # Referred By field contains contractor_code directly
            contractor = code_map.get(inv.referred_by_raw.strip())
            if contractor:
                inv.contractor_id = contractor["id"]
            else:
                # Referred By code not found — downgrade to walk_in
                log.warning(
                    "Invoice %s — Referred By code '%s' not found in contractors",
                    inv.bill_number, inv.referred_by_raw,
                )
                inv.customer_type = CUST_WALK_IN

        elif inv.customer_type == CUST_CONTRACTOR_DIRECT:
            # Particulars field contains contractor_code
            contractor = code_map.get(inv.particulars.strip())
            if contractor:
                inv.contractor_id = contractor["id"]
            else:
                # Particulars code not found — downgrade to walk_in
                log.warning(
                    "Invoice %s — Particulars code '%s' not found in contractors",
                    inv.bill_number, inv.particulars,
                )
                inv.customer_type = CUST_WALK_IN

        # walk_in — no contractor resolution needed